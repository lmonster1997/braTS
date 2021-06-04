import numpy as np
import random
import math
from scipy import stats
import elasticdeform
import torch
import SimpleITK as sitk
import cv2 as cv
from openpyxl import load_workbook


def normalize(slice, bottom=99, down=1):
    """
    normalize image with mean and std for regionnonzero,and clip the value into range
    :param slice:
    :param bottom:
    :param down:
    :return:
    """

    tmp = (slice - torch.min(slice)) / (torch.max(slice) - torch.min(slice))
    tmp = (tmp - torch.mean(tmp)) / torch.std(tmp)

    return tmp

def normalize2(slice, bottom=99, down=1):
    """
    normalize image with mean and std for regionnonzero,and clip the value into range
    :param slice:
    :param bottom:
    :param down:
    :return:
    """
    b = np.percentile(slice, bottom)
    t = np.percentile(slice, down)
    slice = np.clip(slice, t, b)

    image_nonzero = slice[np.nonzero(slice)]
    if np.std(slice) == 0 or np.std(image_nonzero) == 0:
        return slice
    else:
        tmp = (slice - np.min(image_nonzero)) / (np.max(image_nonzero)-np.min(image_nonzero))
        tmp = (tmp - np.mean(tmp)) / np.std(tmp)
        # tmp = (slice - np.mean(slice)) / np.std(slice)
        # since the range of intensities is between 0 and 5000 ,
        # the min in the normalized slice corresponds to 0 intensity in unnormalized slice
        # the min is replaced with -9 just to keep track of 0 intensities
        # # so that we can discard those intensities afterwards when sampling random patches
        # tmp[tmp == tmp.min()] = -9
        return tmp

def load_file_name_list(file_path):
    file_name_list = []
    with open(file_path, 'r') as file_to_read:
        while True:
            lines = file_to_read.readline().strip()  # 整行读取数据
            if not lines:
                break
                pass
            file_name_list.append(lines)
            pass
    return file_name_list

def random_crop_3d(data, mask, crop_size):
    #对图像随机裁剪
    max_D = data.shape[1] - crop_size[0]
    max_H = data.shape[2] - crop_size[1]
    max_W = data.shape[3] - crop_size[2]

    if max_D < 0 or max_H < 0 or max_W < 0:
        return None

    assert data.shape[1]==mask.shape[1], 'The size of data and mask are not equal'
    assert data.shape[2] == mask.shape[2], 'The size of data and mask are not equal'
    assert data.shape[3] == mask.shape[3], 'The size of data and mask are not equal'

    init_D = random.randint(0,max_D)
    init_H = random.randint(0, max_H)
    init_W = random.randint(0, max_W)
    crop_data = data[:,init_D:init_D+crop_size[0],
               init_H:init_H+crop_size[1],init_W:init_W+crop_size[2]]
    crop_mask = mask[:, init_D:init_D + crop_size[0],
               init_H:init_H + crop_size[1], init_W:init_W + crop_size[2]]

    return crop_data, crop_mask

def random_crop_3d2(data, mask, crop_size):
    #对图像随机裁剪
    max_H = data.shape[2] - crop_size[1]
    max_W = data.shape[3] - crop_size[2]

    if max_H < 0 or max_W < 0:
        return None

    assert data.shape[1]==mask.shape[1], 'The size of data and mask are not equal'
    assert data.shape[2] == mask.shape[2], 'The size of data and mask are not equal'
    assert data.shape[3] == mask.shape[3], 'The size of data and mask are not equal'

    min_h = 0
    max_h = data.shape[2]-1
    while(np.sum(mask[0,:,min_h,:])==0):
        min_h+=1
    while (np.sum(mask[0, :, max_h, :]) == 0):
        max_h -= 1

    min_w = 0
    max_w = data.shape[3]-1
    while(np.sum(mask[0,:,:,min_w])==0):
        min_w+=1
    while (np.sum(mask[0, : , :,max_w]) == 0):
        max_w -= 1

    init_H = random.randint(max(0, max_h-128), min(min_h, max_H))
    init_W = random.randint(max(0, max_w-128), min(min_w, max_W))

    crop_data = data[:,:,init_H:init_H+crop_size[1],init_W:init_W+crop_size[2]]
    crop_mask = mask[:, :,init_H:init_H + crop_size[1], init_W:init_W + crop_size[2]]

    return crop_data, crop_mask


def flip3D(X, y):
    # random 3D flipping

    choice = np.random.randint(3)
    # if choice == 0:  # flip on x
    #     X_flip, y_flip = X[:, ::-1, :, :], y[:, ::-1, :, :]
    # if choice == 1:  # flip on y
    #     X_flip, y_flip = X[:, :, ::-1, :], y[:, :, ::-1, :]
    # if choice == 2:  # flip on z
    #     X_flip, y_flip = X[:, :, :, ::-1], y[:, :, :, ::-1]

    if choice == 0:  # flip on x
        X_flip, y_flip = np.flip(X, 1), np.flip(y, 1)
    if choice == 1:  # flip on y
        X_flip, y_flip = np.flip(X, 2), np.flip(y, 2)
    if choice == 2:  # flip on z
        X_flip, y_flip = np.flip(X, 3), np.flip(y, 3)

    return X_flip, y_flip


from scipy.ndimage.interpolation import affine_transform


def rotation3D(X, y):
    # 3D random rotations between 0°-30°
    channel_y = np.shape(y)[0]
    alpha, beta, gamma = np.random.randint(0, 31, size=3) / 180 * np.pi
    Rx = np.array([[1, 0, 0],
                   [0, np.cos(alpha), -np.sin(alpha)],
                   [0, np.sin(alpha), np.cos(alpha)]])

    Ry = np.array([[np.cos(beta), 0, np.sin(beta)],
                   [0, 1, 0],
                   [-np.sin(beta), 0, np.cos(beta)]])

    Rz = np.array([[np.cos(gamma), -np.sin(gamma), 0],
                   [np.sin(gamma), np.cos(gamma), 0],
                   [0, 0, 1]])

    R = np.dot(np.dot(Rx, Ry), Rz)
    X_rot = np.empty_like(X)
    y_rot = np.empty_like(y)

    for channel in range(X.shape[0]):
        X_rot[channel, :, :, :] = affine_transform(X[channel, :, :, :], R, offset=0, order=3, mode='constant')

    for i in range(channel_y):
        y_rot[i,:] = affine_transform(y[i,:], R, offset=0, order=0, mode='constant')

    return X_rot, y_rot


def brightness(X, y):
    """
    Changing the brighness of a image using power-law gamma transformation.
    Gain and gamma are chosen randomly for each image channel.

    Gain chosen between [0.9 - 1.1]
    Gamma chosen between [0.4 - 2.5]

    new_im = gain * im^gamma
    """

    X_new = np.zeros(X.shape)
    for c in range(X.shape[0]):
        im = X[c, :, :, :] - np.min(X[c, :, :, :])  # all positive values

        choice = np.random.random_sample((2,))
        gain = (1.1 - 0.9) * np.random.random_sample() + 0.9
        gamma = (2.5 - 1) * choice[0] + 1

        if choice[1] < 0.5:
            gamma = 1 / gamma

        im_new = gain * (im ** gamma)

        # the mode of new_im is where the background lie, so we must set again these values to 0
        X_new[c, :, :, :] = im_new - stats.mode(im_new, axis=None)[0][0]

    return X_new, y


def elastic(X, y):
    """
    Elastic deformation on a image and its target
    """

    [Xel, yel] = elasticdeform.deform_random_grid([X, y], sigma=5, axis=[(1, 2, 3), (1, 2, 3)], order=[3, 0])

    return Xel, yel

## W

def random_crop_3d_w(data, mask, pred_mask,crop_size):
    #对图像随机裁剪
    max_D = data.shape[1] - crop_size[0]
    max_H = data.shape[2] - crop_size[1]
    max_W = data.shape[3] - crop_size[2]

    if max_D < 0 or max_H < 0 or max_W < 0:
        return None

    assert data.shape[1]==mask.shape[1], 'The size of data and mask are not equal'
    assert data.shape[2] == mask.shape[2], 'The size of data and mask are not equal'
    assert data.shape[3] == mask.shape[3], 'The size of data and mask are not equal'

    init_D = random.randint(0,max_D)
    init_H = random.randint(0, max_H)
    init_W = random.randint(0, max_W)
    crop_data = data[:,init_D:init_D+crop_size[0],
               init_H:init_H+crop_size[1],init_W:init_W+crop_size[2]]
    crop_mask = mask[:, init_D:init_D + crop_size[0],
               init_H:init_H + crop_size[1], init_W:init_W + crop_size[2]]
    crop_pred_mask = pred_mask[:, init_D:init_D + crop_size[0],
                init_H:init_H + crop_size[1], init_W:init_W + crop_size[2]]

    return crop_data, crop_mask, crop_pred_mask


def flip3D_w(X, y,z):
    # random 3D flipping

    choice = np.random.randint(3)

    if choice == 0:  # flip on x
        X_flip, y_flip, z_flip = np.flip(X, 1), np.flip(y, 1), np.flip(z, 3)
    if choice == 1:  # flip on y
        X_flip, y_flip, z_flip = np.flip(X, 2), np.flip(y, 2), np.flip(z, 3)
    if choice == 2:  # flip on z
        X_flip, y_flip, z_flip = np.flip(X, 3), np.flip(y, 3), np.flip(z, 3)

    return X_flip, y_flip, z_flip


from scipy.ndimage.interpolation import affine_transform


def rotation3D_w(X, y,z):
    # 3D random rotations between 0°-30°
    channel_y = np.shape(y)[0]
    alpha, beta, gamma = np.random.randint(0, 31, size=3) / 180 * np.pi
    Rx = np.array([[1, 0, 0],
                   [0, np.cos(alpha), -np.sin(alpha)],
                   [0, np.sin(alpha), np.cos(alpha)]])

    Ry = np.array([[np.cos(beta), 0, np.sin(beta)],
                   [0, 1, 0],
                   [-np.sin(beta), 0, np.cos(beta)]])

    Rz = np.array([[np.cos(gamma), -np.sin(gamma), 0],
                   [np.sin(gamma), np.cos(gamma), 0],
                   [0, 0, 1]])

    R = np.dot(np.dot(Rx, Ry), Rz)
    X_rot = np.empty_like(X)
    y_rot = np.empty_like(y)
    z_rot = np.empty_like(z)

    for channel in range(X.shape[0]):
        X_rot[channel, :, :, :] = affine_transform(X[channel, :, :, :], R, offset=0, order=3, mode='constant')

    for i in range(channel_y):
        y_rot[i,:] = affine_transform(y[i,:], R, offset=0, order=0, mode='constant')
    for i in range(channel_y):
        z_rot[i,:] = affine_transform(z[i,:], R, offset=0, order=0, mode='constant')

    return X_rot, y_rot, z_rot


def brightness_w(X, y,z):
    """
    Changing the brighness of a image using power-law gamma transformation.
    Gain and gamma are chosen randomly for each image channel.

    Gain chosen between [0.9 - 1.1]
    Gamma chosen between [0.4 - 2.5]

    new_im = gain * im^gamma
    """

    X_new = np.zeros(X.shape)
    for c in range(X.shape[0]):
        im = X[c, :, :, :] - np.min(X[c, :, :, :])  # all positive values

        choice = np.random.random_sample((2,))
        gain = (1.1 - 0.9) * np.random.random_sample() + 0.9
        gamma = (2.5 - 1) * choice[0] + 1

        if choice[1] < 0.5:
            gamma = 1 / gamma

        im_new = gain * (im ** gamma)

        # the mode of new_im is where the background lie, so we must set again these values to 0
        X_new[c, :, :, :] = im_new - stats.mode(im_new, axis=None)[0][0]

    return X_new, y


def elastic_w(X, y,z):
    """
    Elastic deformation on a image and its target
    """

    [Xel, yel, zel] = elasticdeform.deform_random_grid([X, y,z], sigma=5, axis=[(1, 2, 3), (1, 2, 3)], order=[3, 0])

    return Xel, yel, zel

def adjust_learning_rate(optimizer, epoch, args, lr_god):
    """Sets the learning rate to the initial LR decayed by epoch"""
    lr = lr_god * ((1 - (epoch // args.epochs))**0.9)
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr

def set_requires_grad(nets, requires_grad=False):
    """Set requies_grad=Fasle for all the networks to avoid unnecessary computations
    Parameters:
        nets (network list)   -- a list of networks
        requires_grad (bool)  -- whether the networks require gradients or not
    """
    if not isinstance(nets, list):
        nets = [nets]
    for net in nets:
        if net is not None:
            for param in net.parameters():
                param.requires_grad = requires_grad


# val
def mask2class(mask):
    mask = torch.unsqueeze(torch.argmax(mask, dim=1), dim=1)
    m = mask.shape[0]
    shape1 = mask.shape[3]
    shape2 = mask.shape[4]
    shape = (m,1,128,shape1,shape2)
    WT = torch.zeros(shape)
    WT[mask == 1] = 1
    WT[mask == 2] = 1
    WT[mask == 3] = 1
    TC = torch.zeros(shape)
    TC[mask == 1] = 1
    TC[mask == 3] = 1
    ET = torch.zeros(shape)
    ET[mask == 3] = 1


    return torch.cat((WT, TC, ET), dim=1)

def mask2class2(mask):
    print('1:',np.sum(mask==1),'2:',np.sum(mask==2),'3:',np.sum(mask==3),)
    shape1 = mask.shape[2]
    shape2 = mask.shape[3]
    shape = (1,128,shape1,shape2)
    WT = np.zeros(shape)
    WT[mask == 1] = 1
    WT[mask == 2] = 1
    WT[mask == 3] = 1
    TC = np.zeros(shape)
    TC[mask == 1] = 1
    TC[mask == 3] = 1
    ET = np.zeros(shape)
    ET[mask == 3] = 1
    print('WT:', np.sum(WT), 'TC:', np.sum(TC), 'ET:', np.sum(ET), )
    return np.concatenate((WT, TC, ET), axis=0)

def class2tri(gt):
    # gt = torch.nn.functional.one_hot(torch.from_numpy(np.array(gt[0].copy(), dtype=int)), 4)
    # gt = np.transpose(gt, (3, 0, 1, 2))

    wt = gt[:, 1, :] + gt[:, 2, :] + gt[:, 3, :]
    tc = gt[:, 1, :] + gt[:, 3, :]
    et = gt[:, 3, :]

    wt = torch.unsqueeze(wt,dim=1)
    tc = torch.unsqueeze(tc, dim=1)
    et = torch.unsqueeze(et, dim=1)

    gt3 = torch.cat((wt, tc, et), dim=1)

    return gt3

def class2mask(cls, shape=(128, 160, 192)):
    cls[np.where(cls>0.8)]=1

    mask = np.zeros(shape)
    ET = cls[0,2,:]
    TC = cls[0,1,:]
    WT = cls[0,0,:]
    mask[WT==1] = 2
    mask[TC == 1] = 1
    mask[ET == 1] = 4

    return mask

def c2logis(classes):
    classes = classes.long()
    logis = torch.zeros((1,2,128,128,128))
    logis[:,0,:] = 1-classes
    logis[:, 1, :] = classes

    return logis

def readnii(mask_path):
    mri = sitk.ReadImage(mask_path)
    ori = mri.GetOrigin()
    mri_array = sitk.GetArrayFromImage(mri)
    mri_array = mri_array[8: 136, 40: 200, 34: 226]
    out_real = sitk.GetImageFromArray(mri_array)
    out_real.SetOrigin(ori)

    return out_real, ori

def centroid(roi):
    bottom = 128
    top = 0
    for i in range(0,128):
        if 1 in roi[i,:]:
            if bottom > i:
                bottom = i
            if top < i:
                top = i
    # print(bottom,top)
    middle = (bottom+top)//2

    img = roi[middle,:]
    M = cv.moments(img)
    # print(M)
    cx = int(M['m10'] / M['m00'])
    cy = int(M['m01'] / M['m00'])

    return cx, cy, middle

def name2label(name):
    label_c = np.ones((np.size(name)))
    for id, n in enumerate(name):
        if(int(n[-3:])>=260 and int(n[-3:])<=335):
            label_c[id] = 0
    return label_c

def read_File(inputpath, sheet):
    wb = load_workbook(inputpath)
    ws = wb[sheet]
    feature = []
    for row in ws.iter_rows():
        feature.append(row)
    m = ws.max_row
    n = ws.max_column
    print(m, n)
    Feature = np.zeros((m,n)).astype(str)
    for i in range(0, m):
        for j in range(0, n):
            Feature[i,j] = feature[i][j].value
    return Feature

def concat_mask(mask1,mask2, mask3, mask4):
    mask = torch.zeros((1,4,128,192,160)).cuda()
    mask[:,:,:, 0:128, 0:128] = mask1
    mask[:, :, :, 64:192, 0:128] = mask2 + mask[:, :, :, 64:192, 0:128]
    mask[:, :, :, 0:128, 32:160] = mask3 + mask[:, :, :, 0:128, 32:160]
    mask[:, :, :, 64:192, 32:160] = mask4 + mask[:, :, :, 64:192, 32:160]
    return mask

def judge(cls):
    if cls.cpu()[0,0]>=0.5:
        return 1
    else:
        return 0

def os2c(os):
    c = []
    for o in os:
        if o<300:
            c.append(0)
        elif o<=450 and o>=300:
            c.append(1)
        elif o>450:
            c.append(2)
        else:
            print('false')

    return c

import os

def pathexists(path):
    if not os.path.exists(path):
        os.mkdir(path)

    return path

if __name__ == '__main__':
    data = np.random.rand(4,155,240,240)
    mask = np.random.rand(1, 155, 240, 240)
    crop_data, crop_mask = rotation3D(data, mask)
    print(crop_data.shape, crop_mask.shape)